"""WhatsApp Message Sender + Smart Data Extractor.

Use this bot only to message contacts who have agreed to hear from you.
The first run displays a QR code for manual WhatsApp Web login.
"""

import json
import random
import re
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from playwright.sync_api import Page, TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright


PROJECT_DIR = Path(__file__).resolve().parent
CONTACTS_FILE = PROJECT_DIR / "contacts.xlsx"
PROFILE_DIR = PROJECT_DIR / "whatsapp_profile"
SCREENSHOT_DIR = PROJECT_DIR / "screenshots"

DEFAULT_MESSAGE = "Hello {name}, this is your daily update."
LOGIN_TIMEOUT = 300_000       # Five minutes for QR login
ELEMENT_TIMEOUT = 20_000


def human_pause(page: Page) -> None:
    """Wait between two and five seconds between important actions."""
    seconds = random.uniform(2, 5)
    page.wait_for_timeout(int(seconds * 1000))


def safe_filename(text: str) -> str:
    """Convert a contact name into a valid screenshot filename."""
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", text).strip("_")
    return cleaned or "contact"


def read_contacts() -> list[dict[str, str]]:
    """Read and validate Name, Phone, and Message from contacts.xlsx."""
    if not CONTACTS_FILE.exists():
        raise FileNotFoundError(
            f"{CONTACTS_FILE.name} was not found. Create it with the columns "
            "Name, Phone, and Message."
        )

    workbook = load_workbook(CONTACTS_FILE, data_only=True)
    sheet = workbook.active
    headers = {
        str(cell.value).strip().lower(): index
        for index, cell in enumerate(sheet[1])
        if cell.value is not None
    }

    missing = {"name", "phone", "message"} - set(headers)
    if missing:
        raise ValueError(
            "contacts.xlsx is missing column(s): " + ", ".join(sorted(missing))
        )

    contacts = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        name_value = row[headers["name"]]
        phone_value = row[headers["phone"]]
        message_value = row[headers["message"]]

        name = "" if name_value is None else str(name_value).strip()
        phone = "" if phone_value is None else str(phone_value).strip()
        message = "" if message_value is None else str(message_value).strip()

        if not name and not phone and not message:
            continue
        if not name or not phone:
            print(f"Skipping row {row_number}: Name and Phone are required.")
            continue

        contacts.append({"name": name, "phone": phone, "message": message})

    if not contacts:
        raise ValueError("contacts.xlsx does not contain any valid contacts.")
    return contacts


def wait_for_login(page: Page) -> None:
    """Wait until the user has scanned the QR code and the chat panel appears."""
    print("Waiting for WhatsApp Web login...")
    print("Scan the QR code with your phone if this is the first run.")
    page.wait_for_selector("#side", state="visible", timeout=LOGIN_TIMEOUT)
    print("WhatsApp Web is ready.")


def get_search_box(page: Page):
    """Find the search textbox inside WhatsApp's left side panel."""
    # In the newer WhatsApp layout this is initially a clickable bar. Clicking
    # it creates/activates the real textbox used for typing the search query.
    search_triggers = [
        page.get_by_text("Search or start a new chat", exact=True),
        page.locator("[aria-label='Search or start a new chat']"),
        page.locator("[title='Search or start a new chat']"),
    ]
    for trigger in search_triggers:
        if trigger.count() and trigger.first.is_visible():
            trigger.first.click()
            page.wait_for_timeout(500)
            break

    candidates = [
        page.get_by_role("textbox", name=re.compile("search", re.IGNORECASE)),
        page.locator("input[placeholder='Search or start a new chat']"),
        page.locator("[contenteditable='true'][aria-placeholder='Search']"),
        page.locator("div[contenteditable='true'][data-tab='3']"),
        page.locator("div[contenteditable='true'][aria-placeholder*='Search']"),
        page.locator("[role='textbox'][aria-label*='Search']"),
        page.locator("input[placeholder*='Search']"),
        page.locator("#side [contenteditable='true']"),
    ]

    for candidate in candidates:
        box = candidate.first
        try:
            box.wait_for(state="visible", timeout=3_000)
            return box
        except PlaywrightTimeoutError:
            continue

    # Final fallback for WhatsApp versions whose search box has no useful label.
    # Ignore the chat footer and choose a visible textbox near the upper-left.
    all_boxes = page.locator("input, [contenteditable='true'][role='textbox']")
    for index in range(all_boxes.count()):
        box = all_boxes.nth(index)
        bounds = box.bounding_box() if box.is_visible() else None
        outside_footer = box.evaluate("element => !element.closest('footer')")
        if bounds and outside_footer and bounds["x"] < 650 and bounds["y"] < 300:
            return box

    raise RuntimeError("WhatsApp search box was not found.")


def get_message_box(page: Page, timeout: int = ELEMENT_TIMEOUT):
    """Find the Type a message textbox in the active chat footer."""
    selectors = [
        "footer div[contenteditable='true'][role='textbox']",
        "footer div[contenteditable='true']",
    ]
    for selector in selectors:
        box = page.locator(selector).first
        try:
            box.wait_for(state="visible", timeout=timeout // len(selectors))
            return box
        except PlaywrightTimeoutError:
            continue
    raise RuntimeError("The 'Type a message' box was not found.")


def search_and_open_chat(page: Page, name: str, phone: str) -> None:
    """Search by phone/name, with a direct-number fallback for unsaved contacts."""
    search_box = get_search_box(page)
    search_value = phone or name

    search_box.click()
    # WhatsApp rebuilds the search element while text is entered. Use the page
    # keyboard after focusing it so later keys do not depend on a stale locator.
    page.keyboard.press("Control+A")
    page.keyboard.press("Backspace")
    page.keyboard.type(search_value, delay=80)
    human_pause(page)

    no_results = page.get_by_text(
        re.compile(r"no chats, contacts or messages found", re.IGNORECASE)
    )
    if no_results.count() == 0:
        # New WhatsApp layouts require selecting a result before Enter.
        page.keyboard.press("ArrowDown")
        page.keyboard.press("Enter")
        try:
            get_message_box(page, timeout=10_000)
            return
        except RuntimeError:
            pass

        # If keyboard selection is unsupported, click the first chat result.
        result_candidates = [
            page.get_by_text(name, exact=True),
            page.locator("#pane-side [role='listitem']"),
            page.locator("#pane-side [role='row']"),
            page.locator("[aria-label*='Search results'] [role='listitem']"),
        ]
        for candidate in result_candidates:
            if candidate.count() and candidate.first.is_visible():
                candidate.first.click()
                try:
                    get_message_box(page, timeout=10_000)
                    return
                except RuntimeError:
                    continue

    # An unsaved but valid WhatsApp number may not appear in normal search.
    digits = re.sub(r"\D", "", phone)
    if not digits:
        raise RuntimeError(f"Contact not found: {name}")
    if not phone.strip().startswith("+"):
        raise RuntimeError(
            f"Phone number for {name} must include the country code, for example "
            "+918939107333. Format the Excel Phone cell as Text."
        )

    print(f"Normal search did not open {name}; trying the phone number.")
    page.goto(f"https://web.whatsapp.com/send?phone={quote(digits)}")
    try:
        get_message_box(page, timeout=ELEMENT_TIMEOUT)
    except RuntimeError as exc:
        raise RuntimeError(f"Contact or WhatsApp number not found: {name}") from exc


def send_message(page: Page, message: str) -> None:
    """Send a message and confirm that it left the message composer."""
    message_box = get_message_box(page)
    matching_text = page.get_by_text(message, exact=True)
    text_count_before = matching_text.count()

    message_box.click()
    message_box.fill(message)
    human_pause(page)
    message_box.press("Enter")

    # WhatsApp changes its outgoing-message CSS classes frequently. Confirm the
    # send using visible text first, then the cleared composer as a fallback.
    for attempt in range(30):
        if page.get_by_text(message, exact=True).count() > text_count_before:
            return

        composer = page.locator("footer [contenteditable='true'][role='textbox']").first
        if composer.count() and composer.is_visible():
            composer_text = " ".join(composer.inner_text().split())
            if not composer_text and attempt >= 4:
                return

        page.wait_for_timeout(500)

    raise RuntimeError("The sent message could not be confirmed.")


def extract_last_three_received(page: Page) -> list[str]:
    """Extract the last three incoming message bubbles from the open chat."""
    incoming = page.locator("div.message-in")
    messages = []

    for index in range(max(0, incoming.count() - 3), incoming.count()):
        text = " ".join(incoming.nth(index).inner_text().split())
        if text:
            messages.append(text)
    return messages[-3:]


def save_reports(results: list[dict], date_stamp: str) -> tuple[Path, Path]:
    """Save the full JSON report and an Excel summary."""
    json_path = PROJECT_DIR / f"whatsapp_report_{date_stamp}.json"
    excel_path = PROJECT_DIR / f"whatsapp_report_{date_stamp}.xlsx"

    with json_path.open("w", encoding="utf-8") as file:
        json.dump(results, file, indent=2, ensure_ascii=False)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "WhatsApp Report"
    sheet.append(
        [
            "Name",
            "Phone",
            "Personalized Message",
            "Status",
            "Sent At",
            "Last 3 Received Messages",
            "Screenshot",
            "Error",
        ]
    )

    for item in results:
        sheet.append(
            [
                item["name"],
                item["phone"],
                item["personalized_message"],
                item["status"],
                item["sent_at"],
                " | ".join(item["last_three_received"]),
                item["screenshot"],
                item["error"],
            ]
        )

    sheet.freeze_panes = "A2"
    widths = {"A": 20, "B": 20, "C": 45, "D": 14, "E": 22, "F": 60, "G": 35, "H": 45}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    workbook.save(excel_path)
    return json_path, excel_path


def main() -> None:
    contacts = read_contacts()
    date_stamp = datetime.now().strftime("%Y-%m-%d")
    run_screenshots = SCREENSHOT_DIR / date_stamp
    run_screenshots.mkdir(parents=True, exist_ok=True)
    results = []

    print(f"Loaded {len(contacts)} contact(s) from contacts.xlsx.")
    confirmation = input("Type SEND to begin messaging consenting contacts: ").strip()
    if confirmation != "SEND":
        print("Cancelled. No messages were sent.")
        return

    with sync_playwright() as playwright:
        context = playwright.chromium.launch_persistent_context(
            user_data_dir=str(PROFILE_DIR),
            headless=False,
            viewport={"width": 1440, "height": 900},
        )
        page = context.pages[0] if context.pages else context.new_page()
        page.goto("https://web.whatsapp.com")

        try:
            wait_for_login(page)

            for number, contact in enumerate(contacts, start=1):
                name = contact["name"]
                template = contact["message"] or DEFAULT_MESSAGE
                personalized = template.replace("{name}", name)
                screenshot_path = run_screenshots / (
                    f"{number:03d}_{safe_filename(name)}.png"
                )

                result = {
                    "name": name,
                    "phone": contact["phone"],
                    "personalized_message": personalized,
                    "status": "Failed",
                    "sent_at": "",
                    "last_three_received": [],
                    "screenshot": "",
                    "error": "",
                }

                print(f"[{number}/{len(contacts)}] Processing {name}...")
                try:
                    search_and_open_chat(page, name, contact["phone"])
                    human_pause(page)
                    send_message(page, personalized)
                    result["status"] = "Sent"
                    result["sent_at"] = datetime.now().isoformat(timespec="seconds")
                    human_pause(page)
                    result["last_three_received"] = extract_last_three_received(page)
                    page.screenshot(path=str(screenshot_path))
                    result["screenshot"] = str(screenshot_path.relative_to(PROJECT_DIR))
                    print(f"Message sent to {name}.")
                except Exception as exc:
                    result["error"] = str(exc)
                    print(f"Could not process {name}: {exc}")

                results.append(result)
                human_pause(page)
        finally:
            context.close()

    json_path, excel_path = save_reports(results, date_stamp)
    print("Automation complete.")
    print(f"JSON report:  {json_path}")
    print(f"Excel report: {excel_path}")


if __name__ == "__main__":
    main()
